# Import pandas for data handling and analysis
import pandas as pd

# Import Path for cleaner file path management
from pathlib import Path

# Import Excel formatting tools
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =========================================================
# FILE PATHS
# =========================================================

# Define the location of the raw input file
RAW_FILE = Path("data/raw/sample_inventory_data.csv")

# Define processed data folder
PROCESSED_DIR = Path("data/processed")

# Define reports folder
REPORTS_DIR = Path("reports")

# Define output file paths
CLEAN_FILE = PROCESSED_DIR / "cleaned_inventory_data.csv"
EXCEPTION_FILE = PROCESSED_DIR / "exception_inventory_data.csv"
REPORT_FILE = REPORTS_DIR / "inventory_report.xlsx"


# =========================================================
# REQUIRED COLUMN VALIDATION
# =========================================================

# Define the columns that must exist in the source data
required_columns = [
    "serial_number",
    "part_number",
    "overall_length",
    "diameter",
    "cal_alpha",
    "cal_od",
    "status",
    "inspection_date",
]


# =========================================================
# LOAD RAW DATA
# =========================================================

# Read the CSV file into a pandas DataFrame
df = pd.read_csv(RAW_FILE)


# =========================================================
# CHECK FOR MISSING REQUIRED COLUMNS
# =========================================================

# Identify any required columns missing from the dataset
missing_columns = [
    col for col in required_columns
    if col not in df.columns
]

# Stop execution if required columns are missing
if missing_columns:
    raise ValueError(
        f"Missing required columns: {missing_columns}"
    )


# =========================================================
# DATA CLEANING
# =========================================================

# Convert inspection_date column into datetime format
# Invalid dates will become NaT (Not a Time)
df["inspection_date"] = pd.to_datetime(
    df["inspection_date"],
    errors="coerce"
)


# =========================================================
# QUALITY VALIDATION
# =========================================================

# Create a blank column to store quality issues
df["quality_issue"] = ""

# Flag records missing CAL_ALPHA
df.loc[
    df["cal_alpha"].isna(),
    "quality_issue"
] += "Missing CAL_ALPHA; "

# Flag records missing CAL_OD
df.loc[
    df["cal_od"].isna(),
    "quality_issue"
] += "Missing CAL_OD; "


# =========================================================
# SPLIT CLEAN VS EXCEPTION RECORDS
# =========================================================

# Clean records contain no quality issues
clean_df = df[df["quality_issue"] == ""]

# Exception records contain at least one quality issue
exception_df = df[df["quality_issue"] != ""]


# =========================================================
# CREATE SUMMARY TABLE
# =========================================================

# Build a high-level metrics summary DataFrame
summary_df = pd.DataFrame({
    "Metric": [
        "Total Records",
        "Clean Records",
        "Exception Records",
        "Unique Part Numbers",
    ],

    "Value": [
        len(df),
        len(clean_df),
        len(exception_df),
        df["part_number"].nunique(),
    ],
})


# =========================================================
# CREATE PART-LEVEL SUMMARY
# =========================================================

# Group data by part number and calculate metrics
part_summary_df = (
    df.groupby("part_number")
    .agg(
        record_count=("serial_number", "count"),

        avg_overall_length=(
            "overall_length",
            "mean"
        ),

        avg_diameter=(
            "diameter",
            "mean"
        ),

        avg_cal_alpha=(
            "cal_alpha",
            "mean"
        ),

        avg_cal_od=(
            "cal_od",
            "mean"
        ),
    )
    .reset_index()
)


# =========================================================
# CREATE OUTPUT DIRECTORIES
# =========================================================

# Create processed directory if it does not exist
PROCESSED_DIR.mkdir(
    parents=True,
    exist_ok=True
)

# Create reports directory if it does not exist
REPORTS_DIR.mkdir(
    parents=True,
    exist_ok=True
)


# =========================================================
# EXPORT CSV FILES
# =========================================================

# Export clean records
clean_df.to_csv(
    CLEAN_FILE,
    index=False
)

# Export exception records
exception_df.to_csv(
    EXCEPTION_FILE,
    index=False
)


# =========================================================
# CREATE EXCEL REPORT
# =========================================================

# Open an Excel writer using openpyxl
with pd.ExcelWriter(
    REPORT_FILE,
    engine="openpyxl"
) as writer:

    # Export summary sheet
    summary_df.to_excel(
        writer,
        sheet_name="Summary",
        index=False
    )

    # Export part summary sheet
    part_summary_df.to_excel(
        writer,
        sheet_name="Part Summary",
        index=False
    )

    # Export clean data sheet
    clean_df.to_excel(
        writer,
        sheet_name="Clean Data",
        index=False
    )

    # Export exception data sheet
    exception_df.to_excel(
        writer,
        sheet_name="Exceptions",
        index=False
    )

    # Access workbook object
    workbook = writer.book

    # Loop through every worksheet
    for sheet_name in workbook.sheetnames:

        # Select worksheet
        worksheet = workbook[sheet_name]

        # Freeze the top row
        worksheet.freeze_panes = "A2"

        # Make header row bold
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        # Automatically adjust column widths
        for column_cells in worksheet.columns:

            # Track maximum cell length
            max_length = 0

            # Get Excel column letter
            column_letter = get_column_letter(
                column_cells[0].column
            )

            # Check each cell in the column
            for cell in column_cells:

                try:
                    cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length

                except:
                    pass

            # Add padding to width
            adjusted_width = max_length + 2

            # Apply width to column
            worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width


# =========================================================
# FINAL STATUS MESSAGES
# =========================================================

print("Clean data exported.")
print("Exception data exported.")
print("Excel report created.")